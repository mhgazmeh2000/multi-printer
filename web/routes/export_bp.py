"""
خروجی داده: CSV, Excel (لاگ و وضعیت پرینترها) – با پشتیبانی از paper_size
"""
import csv
import io
from datetime import datetime
from flask import Blueprint, jsonify, request, make_response, send_file
from flask_login import current_user
from core import store
from core.database import get_log, get_all_logs, get_reset_history
from web.auth import allowed_printer_ips, user_allowed_offices

bp = Blueprint("export", __name__)


# ─── export لاگ با بازه تاریخ ───────────────────────────────────
@bp.route('/api/export/logs')
def export_logs():
    start     = request.args.get('start')
    end       = request.args.get('end')
    fmt       = request.args.get('format', 'csv').lower()
    ip_filter = request.args.get('ip')
    allowed_offices = user_allowed_offices(current_user)
    allowed_ips = allowed_printer_ips(current_user)
    if allowed_offices:
        if ip_filter and ip_filter not in allowed_ips:
            return jsonify({"error": "forbidden"}), 403
        logs = get_all_logs(start, end, 1_000_000, ips=allowed_ips, ip=ip_filter if ip_filter in allowed_ips else None)
    else:
        logs = get_all_logs(start, end, 1_000_000, ip_filter)

    if fmt == 'csv':
        return _csv_logs(logs, start, end)
    elif fmt == 'excel':
        return _excel_logs(logs, start, end)
    return jsonify({"error": "unsupported format"}), 400


def _csv_logs(logs, start, end):
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(['Timestamp','Printer IP','Printer Name','Type','Message',
                'Pages','Color','Code','Severity','Paper Size','Details'])   # ← اضافه شد
    import json
    for ev in logs:
        w.writerow([
            ev.get('timestamp',''), ev.get('printer_ip',''), ev.get('printer_name',''),
            ev.get('type',''), ev.get('message',''), ev.get('pages',''),
            ev.get('color',''), ev.get('code',''), ev.get('severity',''),
            ev.get('paper_size',''),   # ← اضافه شد
            json.dumps({k:v for k,v in ev.items()
                        if k not in ('timestamp','printer_ip','printer_name',
                                     'type','message','pages','color','code','severity','paper_size')},
                       ensure_ascii=False),
        ])
    output.seek(0)
    fname = f"logs_{start or 'begin'}_{end or 'end'}.csv"
    return make_response(output.read(), 200, {
        'Content-Type': 'text/csv; charset=utf-8-sig',
        'Content-Disposition': f'attachment; filename="{fname}"',
    })


def _excel_logs(logs, start, end):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    import json
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Logs"; ws.sheet_view.rightToLeft = True
    thin = Side(style='thin', color='2A3248')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ['زمان','IP','نام دستگاه','نوع','پیام','صفحات','رنگ','کد','اهمیت','سایز کاغذ','جزئیات']   # ← اضافه شد
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font      = Font(bold=True, color='00D4FF', name='Arial', size=10)
        cell.fill      = PatternFill('solid', start_color='0E1018')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = bdr
    SEV_COLORS = {'error':'FF3D3D','warning':'FFD740','success':'00E676','info':'7B86A0'}
    for row, ev in enumerate(logs, 2):
        bg  = '141720' if row % 2 == 0 else '0E1018'
        sev = ev.get('severity','info')
        vals = [
            (ev.get('timestamp','')[:19].replace('T',' ')),
            ev.get('printer_ip',''), ev.get('printer_name',''), ev.get('type',''),
            ev.get('message',''), ev.get('pages',''), ev.get('color',''), ev.get('code',''),
            sev.upper(),
            ev.get('paper_size',''),   # ← اضافه شد
            json.dumps({k:v for k,v in ev.items()
                        if k not in ('timestamp','printer_ip','printer_name',
                                     'type','message','pages','color','code','severity','paper_size')},
                       ensure_ascii=False),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row, col, val)
            cell.font      = Font(name='Arial', size=9, color='E4E8F0')
            cell.fill      = PatternFill('solid', start_color=bg)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = bdr
            if col == 9:
                cell.font = Font(name='Arial', size=9, bold=True,
                                 color=SEV_COLORS.get(sev,'E4E8F0'))
                cell.fill = PatternFill('solid', start_color=bg)
    for i in range(1, 12):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"logs_{start or 'begin'}_{end or 'end'}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── گزارش کامل Excel ───────────────────────────────────────────
@bp.route('/api/export/excel')
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        HAVE_OX = True
    except ImportError:
        HAVE_OX = False

    # ─── دریافت snapshot و اعمال فیلترهای دسترسی ─────────────────
    allowed_offices = user_allowed_offices(current_user)
    allowed_ips = allowed_printer_ips(current_user)
    with store.data_lock:
        snap = list(store.printer_data.values())
    with store.printers_lock:
        cfg = list(store.PRINTERS)
    if allowed_offices:
        snap = [d for d in snap if d.get("ip") in allowed_ips]
        cfg = [p for p in cfg if p.get("ip") in allowed_ips]

    # اگر openpyxl نصب نباشد، fallback به CSV
    if not HAVE_OX:
        return _csv_full_report(snap, cfg)

    wb   = openpyxl.Workbook()
    thin = Side(style='thin', color='2A3248')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hcell(ws, r, c, v, fg="00D4FF", bg="0E1018"):
        cell = ws.cell(r, c, v)
        cell.font      = Font(bold=True, color=fg, name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = bdr
        return cell

    def dcell(ws, r, c, v, bg="141720", fg="E4E8F0"):
        cell = ws.cell(r, c, v)
        cell.font      = Font(name="Arial", size=9, color=fg)
        cell.fill      = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bdr
        return cell

    ROWS = ["141720", "0E1018"]

    # Sheet 1: Status (بدون تغییر)
    ws1 = wb.active; ws1.title = "Printer Status"; ws1.sheet_view.rightToLeft = True
    H1  = ["IP","نام","آنلاین","مدل","سریال","Uptime","کل","رنگی","BW","Copy","Fax","Scan",
           "A4","A3","A4R","A5","B4","T1%","T2%","T3%",
           "Cyan%","Cyan Dots","Magenta%","M Dots","Yellow%","Y Dots","Black%","K Dots","هشدار"]
    for ci, h in enumerate(H1, 1): hcell(ws1, 1, ci, h)
    ws1.row_dimensions[1].height = 30
    for ri, d in enumerate(snap, 2):
        c = d.get("counters",{}); pz = d.get("paper_sizes",{})
        tr= d.get("trays",[]); to = d.get("toners",{})
        dev = d.get("device",{}); online = d.get("online")
        bg  = ROWS[ri % 2]
        t1  = next((t["level"] for t in tr if t["name"]=="Tray 1"), "")
        t2  = next((t["level"] for t in tr if t["name"]=="Tray 2"), "")
        t3  = next((t["level"] for t in tr if t["name"]=="Tray 3"), "")
        scan_total = (c.get("scan_fc") or 0) + (c.get("scan_bw") or 0)
        vals = [d["ip"], d["name"], "✓ آنلاین" if online else "✗ آفلاین",
                dev.get("model",""), dev.get("serial",""), dev.get("uptime_str",""),
                c.get("total",0), c.get("full_color",0), c.get("black_white",0),
                c.get("copy",0), c.get("fax",0), scan_total,
                pz.get("A4",{}).get("total",0), pz.get("A3",{}).get("total",0),
                pz.get("A4R",{}).get("total",0), pz.get("A5",{}).get("total",0), pz.get("B4",{}).get("total",0),
                t1, t2, t3,
                to.get("cyan",{}).get("level",0),    to.get("cyan",{}).get("usage",0),
                to.get("magenta",{}).get("level",0), to.get("magenta",{}).get("usage",0),
                to.get("yellow",{}).get("level",0),  to.get("yellow",{}).get("usage",0),
                to.get("black",{}).get("level",0),   to.get("black",{}).get("usage",0),
                " | ".join(a["message"] for a in d.get("alerts",[]))]
        for ci, v in enumerate(vals, 1): dcell(ws1, ri, ci, v, bg)
        fg = "00E676" if online else "FF3D3D"
        ws1.cell(ri, 3).font = Font(name="Arial", size=9, bold=True, color=fg)
        ws1.cell(ri, 3).fill = PatternFill("solid", start_color=bg)
    col_widths = [14,14,9,20,14,10,10,8,8,7,7,8,9,9,9,9,9,6,6,6,7,14,9,14,9,14,7,14,35]
    for i, w in enumerate(col_widths):
        ws1.column_dimensions[get_column_letter(i+1)].width = w

    # Sheet 2: Job Log (با ستون سایز کاغذ)
    ws3 = wb.create_sheet("Job Log"); ws3.sheet_view.rightToLeft = True
    H3  = ["زمان","IP","نام دستگاه","نوع رویداد","پیام","صفحات","رنگ","کد","اهمیت","سایز کاغذ"]   # ← اضافه شد
    for ci, h in enumerate(H3, 1): hcell(ws3, 1, ci, h)
    ws3.row_dimensions[1].height = 28
    all_ev = []
    for p in cfg:
        for e in get_log(p["ip"], limit=1000):
            all_ev.append((p["ip"], p["name"], e))
    all_ev.sort(key=lambda x: x[2].get("timestamp",""), reverse=True)
    SEV = {"error":"FF3D3D","warning":"FFD740","success":"00E676","info":"7B86A0"}
    for ri, (pip, pname, e) in enumerate(all_ev, 2):
        bg  = ROWS[ri % 2]
        ts  = (e.get("timestamp") or "")[:19].replace("T"," ")
        sev = e.get("severity") or "info"
        vals = [ts, pip, pname, e.get("type",""), e.get("message",""),
                e.get("pages",""), e.get("color",""), e.get("code",""), sev.upper(),
                e.get("paper_size","")]   # ← اضافه شد
        for ci, v in enumerate(vals, 1): dcell(ws3, ri, ci, v, bg)
        ws3.cell(ri, 9).font = Font(name="Arial", size=9, bold=True,
                                    color=SEV.get(sev,"E4E8F0"))
        ws3.cell(ri, 9).fill = PatternFill("solid", start_color=bg)
    for i, w in enumerate([19,14,14,12,36,8,12,8,10,12]):
        ws3.column_dimensions[get_column_letter(i+1)].width = w

    # Sheet 3: Reset History
    ws4 = wb.create_sheet("Reset History"); ws4.sheet_view.rightToLeft = True
    H4 = ["زمان تنظیم","IP","نام دستگاه","رنگ","درصد تنظیم‌شده","کل صفحات در لحظه تنظیم",
          "صفحات چاپ‌شده پس از تنظیم","صفحات به ازای هر 1٪"]
    for ci, h in enumerate(H4, 1): hcell(ws4, 1, ci, h)
    ws4.row_dimensions[1].height = 28
    current_totals = {d.get("ip"): (d.get("counters", {}) or {}).get("total") for d in snap if d.get("ip")}
    reset_rows = get_reset_history(ips=[p["ip"] for p in cfg], limit=5000, current_totals=current_totals)
    for ri, row in enumerate(reset_rows, 2):
        bg = ROWS[ri % 2]
        vals = [
            (row.get("timestamp") or "")[:19].replace("T", " "),
            row.get("printer_ip", ""),
            row.get("printer_name", ""),
            row.get("color", ""),
            row.get("set_level", ""),
            row.get("total_pages_at_reset", ""),
            row.get("pages_printed_after_reset", ""),
            row.get("pages_per_1pct", ""),
        ]
        for ci, v in enumerate(vals, 1):
            dcell(ws4, ri, ci, v, bg)
    for i, w in enumerate([19, 14, 18, 12, 12, 18, 20, 18]):
        ws4.column_dimensions[get_column_letter(i+1)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"printer_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _csv_full_report(snap, cfg):
    out = io.StringIO()
    w   = csv.writer(out)
    w.writerow(["Report", datetime.now().isoformat()])
    w.writerow([])
    w.writerow(["IP","Name","Online","Model","Serial","Total","FC","BW","Copy","Fax",
                "A4","A3","T1%","T2%","CyanPct","MagentaPct","YellowPct","BlackPct","Alerts"])
    for d in snap:
        c = d.get("counters",{}); pz = d.get("paper_sizes",{})
        tr= d.get("trays",[]); to = d.get("toners",{})
        dev = d.get("device",{})
        t1 = next((t["level"] for t in tr if t["name"]=="Tray 1"), "")
        t2 = next((t["level"] for t in tr if t["name"]=="Tray 2"), "")
        w.writerow([d["ip"], d["name"], "Y" if d.get("online") else "N",
            dev.get("model",""), dev.get("serial",""),
            c.get("total",0), c.get("full_color",0), c.get("black_white",0),
            c.get("copy",0), c.get("fax",0),
            pz.get("A4",{}).get("total",0), pz.get("A3",{}).get("total",0), t1, t2,
            to.get("cyan",{}).get("level",0), to.get("magenta",{}).get("level",0),
            to.get("yellow",{}).get("level",0), to.get("black",{}).get("level",0),
            " | ".join(a["message"] for a in d.get("alerts",[]))])
    out.seek(0)
    fname = f"printers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return make_response(out.read(), 200, {
        "Content-Type": "text/csv;charset=utf-8-sig",
        "Content-Disposition": f'attachment;filename="{fname}"',
    })